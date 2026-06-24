from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, session
from app.extensions import db
from app.models import Exam, StudentAttempt, Question, QuestionOption, StudentAnswer, CandidateRegistration
from datetime import datetime, timezone, timedelta
import io
import openpyxl
import uuid
from zoneinfo import ZoneInfo
from app.services.email_templates import generate_result_email
from app.utils.auth import admin_required
from app.constants.timezones import TIMEZONE_CHOICES

admin_exams_bp = Blueprint("admin_exams", __name__, url_prefix="/admin/exams")

@admin_exams_bp.before_request
@admin_required
def require_admin():
    pass

@admin_exams_bp.route("/")
def list_exams():
    search = request.args.get('search', '').strip()
    query = Exam.query
    if search:
        query = query.filter(db.or_(Exam.title.ilike(f"%{search}%"), Exam.exam_code.ilike(f"%{search}%")))
    exams = query.order_by(Exam.created_at.desc()).all()
    return render_template("admin/exams/list.html", exams=exams, search=search)

@admin_exams_bp.route("/create", methods=["GET", "POST"])
def create_exam():
    if request.method == "POST":
        title = request.form.get("title")
        exam_code = request.form.get("exam_code")
        description = request.form.get("description")
        duration_minutes = request.form.get("duration_minutes")
        passing_type = request.form.get("passing_type")
        passing_value = request.form.get("passing_value")
        negative_marking_enabled = request.form.get("negative_marking_enabled") == 'on'
        negative_marks = request.form.get("negative_marks", 0)
        show_result_immediately = request.form.get("show_result_immediately") == 'on'
        restrict_to_time_window = request.form.get("restrict_to_time_window") == 'on'
        start_datetime = request.form.get("start_datetime")
        end_datetime = request.form.get("end_datetime")
        is_active = request.form.get("is_active") == 'on'

        instructions = request.form.get("instructions")
        show_question_numbers = request.form.get("show_question_numbers") == 'on'
        allow_question_navigation = request.form.get("allow_question_navigation") == 'on'
        show_progress_bar = request.form.get("show_progress_bar") == 'on'
        auto_submit_on_timeout = request.form.get("auto_submit_on_timeout") == 'on'
        shuffle_questions = request.form.get("shuffle_questions") == 'on'
        require_candidate_registration = request.form.get("require_candidate_registration") == 'on'
        allow_retest_on_failure = request.form.get("allow_retest_on_failure") == 'on'
        max_attempts = int(request.form.get("max_attempts") or 2)

        # Basic Validation
        if not title or not exam_code or not duration_minutes or not passing_type or not passing_value:
            flash("Please fill out all required fields.", "danger")
            return render_template("admin/exams/form.html", action="Create")

        # Check for unique exam_code
        existing_exam = Exam.query.filter_by(exam_code=exam_code).first()
        if existing_exam:
            flash("Exam Code already exists. Please choose a different one.", "danger")
            return render_template("admin/exams/form.html", action="Create")

        timezone_str = request.form.get("timezone", "Asia/Kolkata")
        tz = ZoneInfo(timezone_str)

        # Parse dates if provided and convert to UTC
        start_dt = None
        if start_datetime:
            dt = datetime.fromisoformat(start_datetime)
            start_dt = dt.replace(tzinfo=tz).astimezone(timezone.utc)
            
        end_dt = None
        if end_datetime:
            dt = datetime.fromisoformat(end_datetime)
            end_dt = dt.replace(tzinfo=tz).astimezone(timezone.utc)

        new_exam = Exam(
            title=title,
            exam_code=exam_code,
            description=description,
            duration_minutes=int(duration_minutes),
            passing_type=passing_type,
            passing_value=float(passing_value),
            negative_marking_enabled=negative_marking_enabled,
            negative_marks=float(negative_marks) if negative_marks else 0.0,
            show_result_immediately=show_result_immediately,
            instructions=instructions,
            show_question_numbers=show_question_numbers,
            allow_question_navigation=allow_question_navigation,
            show_progress_bar=show_progress_bar,
            auto_submit_on_timeout=auto_submit_on_timeout,
            shuffle_questions=shuffle_questions,
            require_candidate_registration=require_candidate_registration,
            allow_retest_on_failure=allow_retest_on_failure,
            max_attempts=max_attempts,
            restrict_to_time_window=restrict_to_time_window,
            start_datetime=start_dt,
            end_datetime=end_dt,
            timezone=timezone_str,
            is_active=is_active
        )

        db.session.add(new_exam)
        db.session.commit()
        flash("Exam created successfully.", "success")
        return redirect(url_for("admin_exams.list_exams"))

    return render_template("admin/exams/form.html", action="Create", timezones=TIMEZONE_CHOICES)

@admin_exams_bp.route("/<int:id>/edit", methods=["GET", "POST"])
def edit_exam(id):
    exam = Exam.query.get_or_404(id)
    if request.method == "POST":
        title = request.form.get("title")
        exam_code = request.form.get("exam_code")
        
        # Check uniqueness of exam_code if changed
        if exam_code != exam.exam_code:
            existing_exam = Exam.query.filter_by(exam_code=exam_code).first()
            if existing_exam:
                flash("Exam Code already exists. Please choose a different one.", "danger")
                return render_template("admin/exams/form.html", action="Edit", exam=exam, timezones=TIMEZONE_CHOICES)
            exam.exam_code = exam_code

        exam.title = title
        exam.description = request.form.get("description")
        exam.duration_minutes = int(request.form.get("duration_minutes") or 0)
        exam.passing_type = request.form.get("passing_type")
        exam.passing_value = float(request.form.get("passing_value") or 0.0)
        exam.negative_marking_enabled = request.form.get("negative_marking_enabled") == 'on'
        exam.negative_marks = float(request.form.get("negative_marks") or 0.0)
        exam.show_result_immediately = request.form.get("show_result_immediately") == 'on'
        
        exam.instructions = request.form.get("instructions")
        exam.show_question_numbers = request.form.get("show_question_numbers") == 'on'
        exam.allow_question_navigation = request.form.get("allow_question_navigation") == 'on'
        exam.show_progress_bar = request.form.get("show_progress_bar") == 'on'
        exam.auto_submit_on_timeout = request.form.get("auto_submit_on_timeout") == 'on'
        exam.shuffle_questions = request.form.get("shuffle_questions") == 'on'
        exam.require_candidate_registration = request.form.get("require_candidate_registration") == 'on'
        exam.allow_retest_on_failure = request.form.get("allow_retest_on_failure") == 'on'
        exam.max_attempts = int(request.form.get("max_attempts") or 2)

        exam.restrict_to_time_window = request.form.get("restrict_to_time_window") == 'on'
        
        timezone_str = request.form.get("timezone", "Asia/Kolkata")
        exam.timezone = timezone_str
        tz = ZoneInfo(timezone_str)
        
        start_datetime = request.form.get("start_datetime")
        if start_datetime:
            dt = datetime.fromisoformat(start_datetime)
            exam.start_datetime = dt.replace(tzinfo=tz).astimezone(timezone.utc)
        else:
            exam.start_datetime = None
            
        end_datetime = request.form.get("end_datetime")
        if end_datetime:
            dt = datetime.fromisoformat(end_datetime)
            exam.end_datetime = dt.replace(tzinfo=tz).astimezone(timezone.utc)
        else:
            exam.end_datetime = None
        
        exam.is_active = request.form.get("is_active") == 'on'

        db.session.commit()
        flash("Exam updated successfully.", "success")
        return redirect(url_for("admin_exams.list_exams"))

    attempt_count = StudentAttempt.query.filter_by(exam_id=exam.id).count()
    return render_template("admin/exams/form.html", action="Edit", exam=exam, timezones=TIMEZONE_CHOICES, attempt_count=attempt_count)

@admin_exams_bp.route("/<int:id>/toggle", methods=["POST"])
def toggle_exam(id):
    exam = Exam.query.get_or_404(id)
    exam.is_active = not exam.is_active
    db.session.commit()
    status = "activated" if exam.is_active else "deactivated"
    flash(f"Exam '{exam.title}' has been {status}.", "success")
    return redirect(url_for("admin_exams.list_exams"))

@admin_exams_bp.route("/<int:exam_id>/duplicate", methods=["POST"])
def duplicate_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    
    # Duplicate exam
    new_exam = Exam(
        title=f"Copy of {exam.title}",
        exam_code=f"COPY-{uuid.uuid4().hex[:8].upper()}",
        description=exam.description,
        instructions=exam.instructions,
        duration_minutes=exam.duration_minutes,
        passing_type=exam.passing_type,
        passing_value=exam.passing_value,
        show_question_numbers=exam.show_question_numbers,
        allow_question_navigation=exam.allow_question_navigation,
        show_progress_bar=exam.show_progress_bar,
        auto_submit_on_timeout=exam.auto_submit_on_timeout,
        shuffle_questions=exam.shuffle_questions,
        negative_marking_enabled=exam.negative_marking_enabled,
        negative_marks=exam.negative_marks,
        show_result_immediately=exam.show_result_immediately,
        restrict_to_time_window=exam.restrict_to_time_window,
        start_datetime=exam.start_datetime,
        end_datetime=exam.end_datetime,
        is_active=False
    )
    db.session.add(new_exam)
    db.session.flush()

    # Duplicate questions
    questions = Question.query.filter_by(exam_id=exam.id).all()
    for q in questions:
        new_q = Question(
            exam_id=new_exam.id,
            question_text=q.question_text,
            marks=q.marks,
            display_order=q.display_order
        )
        db.session.add(new_q)
        db.session.flush()

        options = QuestionOption.query.filter_by(question_id=q.id).all()
        for o in options:
            new_o = QuestionOption(
                question_id=new_q.id,
                option_text=o.option_text,
                option_order=o.option_order
            )
            db.session.add(new_o)
            db.session.flush()
            if o.id == q.correct_option_id:
                new_q.correct_option_id = new_o.id
                
    db.session.commit()
    flash(f"Exam duplicated successfully as '{new_exam.title}'.", "success")
    return redirect(url_for('admin_exams.list_exams'))

@admin_exams_bp.route("/<int:exam_id>/delete", methods=["POST"])
def delete_exam(exam_id):
    if not session.get("sensitive_action_verified"):
        session["post_verification_redirect"] = url_for('admin_exams.list_exams')
        return redirect(url_for('admin_backup.auth'))
        
    exam = Exam.query.get_or_404(exam_id)
    
    # 1. Delete StudentAnswers and Attempts
    attempts = StudentAttempt.query.filter_by(exam_id=exam.id).all()
    attempt_ids = [a.id for a in attempts]
    if attempt_ids:
        StudentAnswer.query.filter(StudentAnswer.attempt_id.in_(attempt_ids)).delete(synchronize_session=False)
        StudentAttempt.query.filter_by(exam_id=exam.id).delete(synchronize_session=False)

    # 2. Delete QuestionOptions and Questions
    questions = Question.query.filter_by(exam_id=exam.id).all()
    question_ids = [q.id for q in questions]
    if question_ids:
        QuestionOption.query.filter(QuestionOption.question_id.in_(question_ids)).delete(synchronize_session=False)
        Question.query.filter_by(exam_id=exam.id).delete(synchronize_session=False)

    # 3. Delete Exam
    db.session.delete(exam)
    db.session.commit()
    
    flash("Exam deleted successfully.", "success")
    return redirect(url_for('admin_exams.list_exams'))

@admin_exams_bp.route("/<int:exam_id>/clear-all-attempts", methods=["POST"])
def clear_all_attempts(exam_id):
    from flask import current_app
    exam = Exam.query.get_or_404(exam_id)
    exam_code_input = request.form.get('exam_code_confirm')
    pin_input = request.form.get('security_pin')
    
    if pin_input != current_app.config.get('ADMIN_SECURITY_PIN'):
        flash("Incorrect security PIN. Attempts were not deleted.", "danger")
        return redirect(url_for('admin_exams.edit_exam', id=exam_id))
    
    if exam_code_input != exam.exam_code:
        flash("Incorrect exam code. Attempts were not deleted.", "danger")
        return redirect(url_for('admin_exams.edit_exam', id=exam_id))
        
    try:
        attempts = StudentAttempt.query.filter_by(exam_id=exam.id).all()
        attempt_ids = [a.id for a in attempts]
        
        if attempt_ids:
            StudentAnswer.query.filter(StudentAnswer.attempt_id.in_(attempt_ids)).delete(synchronize_session=False)
            StudentAttempt.query.filter(StudentAttempt.exam_id == exam.id).delete(synchronize_session=False)
            db.session.commit()
            flash(f"Successfully cleared all {len(attempt_ids)} attempts for {exam.title}.", "success")
        else:
            flash("No attempts found to clear.", "info")
            
    except Exception as e:
        db.session.rollback()
        flash("An error occurred while deleting attempts. Please try again.", "danger")
        
    return redirect(url_for('admin_exams.edit_exam', id=exam_id))

def get_filtered_attempts(exam_id):
    query = StudentAttempt.query.filter_by(exam_id=exam_id)
    
    search = request.args.get('search', '').strip()
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    if search:
        query = query.filter(
            db.or_(
                StudentAttempt.student_name.ilike(f"%{search}%"),
                StudentAttempt.student_email.ilike(f"%{search}%")
            )
        )
        
    if from_date_str:
        try:
            from_date = datetime.strptime(from_date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            query = query.filter(StudentAttempt.started_at >= from_date)
        except ValueError:
            pass
            
    if to_date_str:
        try:
            to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
            to_date = to_date + timedelta(days=1) - timedelta(seconds=1)
            to_date = to_date.replace(tzinfo=timezone.utc)
            query = query.filter(StudentAttempt.started_at <= to_date)
        except ValueError:
            pass
            
    return query.order_by(StudentAttempt.submitted_at.desc()).all()

@admin_exams_bp.route("/<int:exam_id>/attempts")
def view_attempts(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    attempts = get_filtered_attempts(exam_id)
    
    # Calculate Attempt #
    all_attempts = StudentAttempt.query.filter_by(exam_id=exam_id).order_by(StudentAttempt.started_at.asc()).all()
    attempt_counters = {}
    attempt_num_map = {}
    for att in all_attempts:
        email = att.student_email
        attempt_counters[email] = attempt_counters.get(email, 0) + 1
        attempt_num_map[att.id] = attempt_counters[email]
        
    for att in attempts:
        att.attempt_number = attempt_num_map.get(att.id, 1)
        
    return render_template("admin/exams/attempts.html", exam=exam, attempts=attempts)

@admin_exams_bp.route("/<int:exam_id>/export-results")
def export_results(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    attempts = get_filtered_attempts(exam_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    
    headers = [
        "Student Name", "Student Email", "Attempt Status", 
        "Marks Obtained", "Percentage Score", "Correct Answers", 
        "Wrong Answers", "Unanswered Questions", "Pass/Fail Status", "Submitted At"
    ]
    ws.append(headers)
    
    for attempt in attempts:
        row = [
            attempt.student_name,
            attempt.student_email,
            attempt.status.title(),
            attempt.total_marks_obtained if attempt.total_marks_obtained is not None else "-",
            f"{attempt.percentage_score:.1f}%" if attempt.percentage_score is not None else "-",
            attempt.correct_count if attempt.correct_count is not None else "-",
            attempt.wrong_count if attempt.wrong_count is not None else "-",
            attempt.unanswered_count if attempt.unanswered_count is not None else "-",
            attempt.result_status if attempt.result_status is not None else "-",
            attempt.submitted_at.replace(tzinfo=timezone.utc).astimezone(ZoneInfo(exam.timezone)).strftime('%B %d, %Y %I:%M %p %Z') if attempt.submitted_at else "-"
        ]
        ws.append(row)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"{exam.exam_code}_results.xlsx"
    return send_file(out, download_name=filename, as_attachment=True)

@admin_exams_bp.route("/<int:exam_id>/export-selected-results", methods=["POST"])
def export_selected_results(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    attempt_ids = request.form.getlist("attempt_ids")
    
    if not attempt_ids:
        flash("Please select at least one attempt.", "warning")
        return redirect(url_for('admin_exams.view_attempts', exam_id=exam.id))
        
    attempts = StudentAttempt.query.filter(
        StudentAttempt.exam_id == exam.id,
        StudentAttempt.id.in_(attempt_ids)
    ).order_by(StudentAttempt.submitted_at.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    
    headers = [
        "Student Name", "Student Email", "Attempt Status", 
        "Marks Obtained", "Percentage Score", "Correct Answers", 
        "Wrong Answers", "Unanswered Questions", "Pass/Fail Status", "Submitted At"
    ]
    ws.append(headers)
    
    for attempt in attempts:
        row = [
            attempt.student_name,
            attempt.student_email,
            attempt.status.title(),
            attempt.total_marks_obtained if attempt.total_marks_obtained is not None else "-",
            f"{attempt.percentage_score:.1f}%" if attempt.percentage_score is not None else "-",
            attempt.correct_count if attempt.correct_count is not None else "-",
            attempt.wrong_count if attempt.wrong_count is not None else "-",
            attempt.unanswered_count if attempt.unanswered_count is not None else "-",
            attempt.result_status if attempt.result_status is not None else "-",
            attempt.submitted_at.replace(tzinfo=timezone.utc).astimezone(ZoneInfo(exam.timezone)).strftime('%B %d, %Y %I:%M %p %Z') if attempt.submitted_at else "-"
        ]
        ws.append(row)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    filename = f"{exam.exam_code}_results.xlsx"
    return send_file(out, download_name=filename, as_attachment=True)

@admin_exams_bp.route("/<int:exam_id>/delete-selected-attempts", methods=["POST"])
def delete_selected_attempts(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    attempt_ids = request.form.getlist("attempt_ids")
    
    if not attempt_ids:
        flash("Please select at least one attempt to delete.", "warning")
        return redirect(url_for('admin_exams.view_attempts', exam_id=exam.id))
        
    try:
        StudentAnswer.query.filter(StudentAnswer.attempt_id.in_(attempt_ids)).delete(synchronize_session=False)
        StudentAttempt.query.filter(
            StudentAttempt.exam_id == exam.id,
            StudentAttempt.id.in_(attempt_ids)
        ).delete(synchronize_session=False)
        
        db.session.commit()
        flash(f"Successfully deleted {len(attempt_ids)} selected attempts.", "success")
    except Exception as e:
        db.session.rollback()
        flash("An error occurred while deleting selected attempts. Please try again.", "danger")
        
    return redirect(url_for('admin_exams.view_attempts', exam_id=exam.id))

@admin_exams_bp.route("/attempts/<int:attempt_id>/compose-email")
def compose_result_email(attempt_id):
    attempt = StudentAttempt.query.get_or_404(attempt_id)
    
    # Calculate total marks from questions
    total_marks = sum([q.marks for q in attempt.exam.questions])
    
    email_data = generate_result_email(attempt, total_marks)
    
    return render_template("admin/attempts/compose_email.html", 
                           attempt=attempt, 
                           email_data=email_data)

# --- CANDIDATE REGISTRATION ROUTES ---

@admin_exams_bp.route("/<int:exam_id>/candidates")
def view_candidates(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    search = request.args.get('search', '').strip()
    
    query = CandidateRegistration.query.filter_by(exam_id=exam.id)
    
    if search:
        query = query.filter(db.or_(
            CandidateRegistration.candidate_name.ilike(f'%{search}%'),
            CandidateRegistration.email.ilike(f'%{search}%')
        ))
        
    candidates = query.order_by(CandidateRegistration.created_at.desc()).all()
    
    # Compute status based on retest logic
    all_attempts = StudentAttempt.query.filter_by(exam_id=exam.id).order_by(StudentAttempt.started_at.asc()).all()
    email_to_attempts = {}
    for att in all_attempts:
        email = att.student_email.lower()
        if email not in email_to_attempts:
            email_to_attempts[email] = []
        email_to_attempts[email].append(att)
        
    stats = {'total': len(candidates), 'passed': 0, 'failed_retest': 0, 'failed_exhausted': 0, 'pending': 0}
    
    for c in candidates:
        email = c.email.lower()
        c_attempts = email_to_attempts.get(email, [])
        
        if not c_attempts:
            c.status_label = "Not Attempted"
            c.status_code = "pending"
            stats['pending'] += 1
        else:
            passed = any(att.result_status == "Pass" for att in c_attempts)
            if passed:
                c.status_label = "Passed"
                c.status_code = "passed"
                stats['passed'] += 1
            else:
                if exam.allow_retest_on_failure and len(c_attempts) < exam.max_attempts:
                    c.status_label = "Failed (Retest Available)"
                    c.status_code = "failed_retest"
                    stats['failed_retest'] += 1
                else:
                    c.status_label = "Failed (Attempts Exhausted)"
                    c.status_code = "failed_exhausted"
                    stats['failed_exhausted'] += 1
            
    return render_template("admin/exams/candidates.html", exam=exam, candidates=candidates, stats=stats)

@admin_exams_bp.route("/<int:exam_id>/candidates/add", methods=["POST"])
def add_candidate(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    name = request.form.get("candidate_name")
    email = request.form.get("email")
    phone = request.form.get("phone")
    
    if not name or not email:
        flash("Name and email are required.", "danger")
        return redirect(url_for('admin_exams.view_candidates', exam_id=exam.id))
        
    existing = CandidateRegistration.query.filter_by(exam_id=exam.id, email=email).first()
    if existing:
        flash(f"Candidate with email {email} is already registered.", "warning")
        return redirect(url_for('admin_exams.view_candidates', exam_id=exam.id))
        
    new_candidate = CandidateRegistration(
        candidate_name=name,
        email=email,
        phone=phone,
        exam_id=exam.id
    )
    db.session.add(new_candidate)
    db.session.commit()
    
    flash("Candidate registered successfully.", "success")
    return redirect(url_for('admin_exams.view_candidates', exam_id=exam.id))

@admin_exams_bp.route("/<int:exam_id>/candidates/<int:candidate_id>/delete", methods=["POST"])
def delete_candidate(exam_id, candidate_id):
    candidate = CandidateRegistration.query.get_or_404(candidate_id)
    if candidate.exam_id != exam_id:
        abort(404)
        
    db.session.delete(candidate)
    db.session.commit()
    flash("Candidate registration deleted.", "success")
    return redirect(url_for('admin_exams.view_candidates', exam_id=exam_id))

@admin_exams_bp.route("/<int:exam_id>/candidates/delete_selected", methods=["POST"])
def delete_selected_candidates(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    candidate_ids = request.form.getlist("candidate_ids")
    
    if not candidate_ids:
        flash("No candidates selected for deletion.", "warning")
        return redirect(url_for('admin_exams.view_candidates', exam_id=exam.id))
        
    try:
        CandidateRegistration.query.filter(
            CandidateRegistration.id.in_(candidate_ids),
            CandidateRegistration.exam_id == exam.id
        ).delete(synchronize_session=False)
        db.session.commit()
        flash(f"Successfully deleted {len(candidate_ids)} selected candidates.", "success")
    except Exception as e:
        db.session.rollback()
        flash("An error occurred while deleting selected candidates.", "danger")
        
    return redirect(url_for('admin_exams.view_candidates', exam_id=exam.id))

@admin_exams_bp.route("/<int:exam_id>/candidates/import", methods=["POST"])
def import_candidates(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    
    if 'file' not in request.files:
        flash('No file part', 'danger')
        return redirect(url_for('admin_exams.view_candidates', exam_id=exam.id))
        
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('admin_exams.view_candidates', exam_id=exam.id))
        
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        flash('Invalid file format. Please upload an Excel file.', 'danger')
        return redirect(url_for('admin_exams.view_candidates', exam_id=exam.id))
        
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        imported = 0
        skipped = 0
        blanks = 0
        
        # Read headers
        headers = [cell.value for cell in sheet[1]]
        
        # Simple mapping assuming Name, Email, Phone
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2:
                name = str(row[0]).strip() if row[0] else ''
                email = str(row[1]).strip() if row[1] else ''
                phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
                
                if not name or not email:
                    blanks += 1
                    continue
                    
                existing = CandidateRegistration.query.filter_by(exam_id=exam.id, email=email).first()
                if existing:
                    skipped += 1
                    continue
                    
                new_candidate = CandidateRegistration(
                    candidate_name=name,
                    email=email,
                    phone=phone,
                    exam_id=exam.id
                )
                db.session.add(new_candidate)
                imported += 1
            else:
                blanks += 1
                
        db.session.commit()
        flash(f"Import complete: {imported} imported, {skipped} duplicate emails skipped, {blanks} invalid/blank rows skipped.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error processing file: {str(e)}", "danger")
        
    return redirect(url_for('admin_exams.view_candidates', exam_id=exam.id))

@admin_exams_bp.route("/<int:exam_id>/candidates/export")
def export_candidates(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    search = request.args.get('search', '').strip()
    
    query = CandidateRegistration.query.filter_by(exam_id=exam.id)
    if search:
        query = query.filter(db.or_(
            CandidateRegistration.candidate_name.ilike(f'%{search}%'),
            CandidateRegistration.email.ilike(f'%{search}%')
        ))
        
    candidates = query.order_by(CandidateRegistration.created_at.desc()).all()
    attempted_emails = {a.student_email.lower() for a in StudentAttempt.query.filter_by(exam_id=exam.id).all()}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"
    
    headers = ["Name", "Email", "Phone", "Status", "Registered On"]
    ws.append(headers)
    
    from app.utils.time_filters import format_datetime_tz
    
    for c in candidates:
        status = "Attempted" if c.email.lower() in attempted_emails else "Pending"
        registered_on = format_datetime_tz(c.created_at, exam.timezone, '%Y-%m-%d %H:%M:%S') if c.created_at else "-"
        ws.append([
            c.candidate_name,
            c.email,
            c.phone or "-",
            status,
            registered_on
        ])
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"{exam.exam_code}_Candidates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@admin_exams_bp.route("/candidates/template")
def download_candidate_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    headers = ["Name", "Email", "Phone"]
    ws.append(headers)
    
    # Add a sample row
    ws.append(["John Doe", "john.doe@example.com", "+1234567890"])
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="candidate_template.xlsx"
    )
